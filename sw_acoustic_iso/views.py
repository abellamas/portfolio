from django.shortcuts import render
from django.http.response import HttpResponse
import plotly.graph_objects as go
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import numpy as np
from sw_acoustic_iso.models import Materials
from sw_acoustic_iso.forms import MaterialsPanelForm, DimensionsPanel
from sw_acoustic_iso.acoustic import Panel
from datetime import datetime
import json


# Create your views here.
def index(request):
    if request.method == "POST":
        material_form = MaterialsPanelForm(request.POST)
        dimensions_form = DimensionsPanel(request.POST)
        if material_form.is_valid() & dimensions_form.is_valid():
            
            material = Materials.objects.get(id=request.POST.get('material'))
            lx = request.POST.get('l_x')
            ly = request.POST.get('l_y')
            thickness = request.POST.get('thickness')
            
            panel = Panel(material.material, material.density, material.young_module, material.loss_factor, material.poisson_module, lx, ly, thickness)
            
            f_per_thirds = [20, 25, 31.5, 40, 50, 63, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000, 6300, 8000, 10000, 12500, 16000, 20000]
            
            f_cremer, r_cremer = panel.cremer_model(f_per_thirds)
            f_davy, r_davy = panel.davy_model(f_per_thirds) 
            f_sharp, r_sharp = panel.sharp_model(f_per_thirds)
            f_iso, r_iso = panel.iso_model(f_per_thirds)
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=f_cremer, y=r_cremer, name="Cremer", line=dict(color='blue', width=1)))
            fig.add_trace(go.Scatter(x=f_davy, y=r_davy, name="Davy", line=dict(color='green', width=1)))
            fig.add_trace(go.Scatter(x=f_sharp, y=r_sharp, name="Sharp", line=dict(color='red', width=1)))
            fig.add_trace(go.Scatter(x=f_iso, y=r_iso, name="ISO 12354-1", line=dict(color='orange', width=1)))
            
            fig.update_xaxes(type="log", tickvals=f_per_thirds, ticktext=f_per_thirds)
            fig.update_layout(
                legend=dict(
                    title="Modelos",
                    orientation="h",
                    # entrywidth=70,
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                ),
                yaxis_title = "R [dB]",
                xaxis_title = "Frecuencia [Hz]",
                template = 'plotly_white'
            )
            
            fig_html = fig.to_html()
            
            reduction_df = pd.DataFrame(data=[r_cremer, r_davy, r_sharp, r_iso], columns=f_per_thirds, index=['Cremer', 'Davy', 'Sharp', 'ISO'])
                
            # Load the workbook
            wb = load_workbook('data/template.xlsx')
            ws = wb.active
            # Convert the dataframe to rows
            rows = dataframe_to_rows(reduction_df, index=False)
            ws['C2'] = material.material
            ws['C3'] = float(lx)
            ws['C4'] = float(ly)
            ws['C5'] = float(thickness)
            ws['C6'] = panel.freq_critic
            # Write the rows to the worksheet
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx+7, column=c_idx+2, value=value)

            # Save the workbook
            wb.save('data/templates.xlsx')
            
            reduction_json = reduction_df.to_json(orient="table")
            reduction_arr = []
            reduction_arr = json.loads(reduction_json)
            print(reduction_arr)
                        
            return render(request, 'base.html', {'material_form': material_form, 'dimensions_form': dimensions_form, 'reduction_arr': reduction_arr, 'fig_html' : fig_html, 'stiffness' : panel.stiffness, 'mass_sup':panel.mass_sup, 'freq_res': panel.freq_res, 'freq_critic': panel.freq_critic, 'freq_density':panel.freq_density})
            
        else:
            print("Error en la validacion")
            
    elif request.method == "GET":
        dimensions_form = DimensionsPanel()
        material_form = MaterialsPanelForm()
        return render(request, 'base.html', {'material_form': material_form, 'dimensions_form': dimensions_form})
    else:
        return render(request, 'base.html')

def export_excel(request):
    hms = datetime.now()
    response = HttpResponse(open("data/templates.xlsx", 'rb').read())
    response['Content-Type'] = 'application/ms-excel'
    response['Content-Disposition'] = f"attachment; filename=aislamiento_ruido_{hms}.xlsx"
    return response 
    
